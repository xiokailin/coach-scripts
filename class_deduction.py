#!/usr/bin/env python3
"""
銷課自動化腳本（範例）

把上次銷課日到今天之間、行事曆上實際上過的課，從公司合約名單（.xlsx）的
「剩餘堂數」欄逐筆扣除。

為什麼不用 openpyxl 直接存檔：
  公司合約單內含大量 charts / drawings，openpyxl 存檔會破壞這些物件。
  因此採「把 xlsx 當成 zip，只改目標工作表 XML 的單一儲存格數值」的做法，
  其餘檔案內容原樣保留。

用法：
  python3 class_deduction.py            # 銷上次到今天
  python3 class_deduction.py --since 2026/6/1   # 指定起算日（覆蓋紀錄檔）
  python3 class_deduction.py --dry-run  # 只試算、不寫回（建議第一次先跑這個）

前置需求：
  - gws（Google Workspace CLI）已授權，可讀 Calendar、讀寫 Drive
  - openpyxl（pip install openpyxl）僅用於「讀回驗證」，不用於存檔
"""

import sys, os, re, json, shutil, zipfile, subprocess
from datetime import date, datetime, timedelta
from pathlib import Path

# ──────────────────────────────────────────
# 設定（請改成你自己的資料）
# ──────────────────────────────────────────

GWS = "/opt/homebrew/bin/gws"

# 公司合約名單 .xlsx 在 Google Drive 的 fileId
# 從檔案的分享網址取得：
#   https://drive.google.com/file/d/【這段就是 fileId】/view
CONTRACT_FILE_ID = "YOUR_CONTRACT_XLSX_FILE_ID"

# 合約名單主分頁在 xlsx 內部對應的 worksheet XML 檔名
# 開 xlsx（當 zip）看 xl/worksheets/ 底下有哪些 sheetN.xml，
# 對照 xl/workbook.xml 的分頁順序找出主分頁是第幾個
CONTRACT_SHEET_XML = "xl/worksheets/sheet4.xml"

# 你在合約名單「教練」欄顯示的姓名
TARGET_TRAINER = "你的姓名"

# 合約名單欄位對應（A=0, B=1, ... 的 1-based 欄字母）
COL_TRAINER = "C"   # 教練
COL_STUDENT = "D"   # 學員姓名
COL_STATUS  = "G"   # 合約狀況（只處理「進行中」）
COL_TOTAL   = "O"   # 總堂數
COL_REMAIN  = "R"   # 剩餘堂數 ← 銷課只改這一欄

# 只處理「進行中」的合約，以下狀況跳過
SKIP_STATUS = ("畢業", "續約", "退費")

# 共用合約：兩人共用同一份合約，一堂只扣一次
# 格式：{"行事曆上會出現的名字": "合約名單上登記的名字"}
# 範例（請替換成你的實際配對）：
SHARED_CONTRACT = {
    # "李大華": "王小明",   # 李大華上課時，扣王小明那筆合約
    # "陳美玲": "陳志明",
}

# 上次銷課日期紀錄檔
LAST_DATE_FILE = Path.home() / ".claude" / "last_deduction_date.json"

# 行事曆上要排除、不算課的關鍵字
EXCLUDE_KEYWORDS = ("自主訓練", "體驗", "請假")
# 算課的關鍵字
COUNT_KEYWORDS = ("一對一", "一對二")


# ──────────────────────────────────────────
# Calendar：累計每位學員應扣堂數
# ──────────────────────────────────────────

def _gws_json(args):
    """執行 gws 並從輸出中解析出 JSON（gws 會在前面印雜訊）。"""
    result = subprocess.run([GWS] + args, capture_output=True, text=True)
    raw = result.stdout
    idx = raw.find("{")
    if idx == -1:
        return {}
    return json.loads(raw[idx:])


def list_events(day):
    """讀取某一天的行事曆行程。"""
    tmin = f"{day:%Y-%m-%d}T00:00:00+08:00"
    tmax = f"{day:%Y-%m-%d}T23:59:59+08:00"
    data = _gws_json([
        "calendar", "events", "list",
        "--params", json.dumps({
            "calendarId": "primary",
            "timeMin": tmin, "timeMax": tmax,
            "singleEvents": True, "orderBy": "startTime",
        })
    ])
    return data.get("items", [])


def collect_deductions(d_start, d_end):
    """從 d_start 到 d_end（含）逐日累計各學員應扣堂數。"""
    counts = {}          # {合約登記名字: 扣堂數}
    not_in_table = set() # 行事曆有、但需提醒補登的名字
    day = d_start
    while day <= d_end:
        for e in list_events(day):
            summary = e.get("summary", "")
            if e.get("status") == "cancelled":
                continue
            if any(k in summary for k in EXCLUDE_KEYWORDS):
                continue
            if not any(k in summary for k in COUNT_KEYWORDS):
                continue
            # 從標題抓學員名字（去掉「一對一」等課程字樣）
            name = re.sub(r"(一對一|一對二).*$", "", summary).strip()
            if not name:
                continue
            # 共用合約：換成合約登記名字
            name = SHARED_CONTRACT.get(name, name)
            counts[name] = counts.get(name, 0) + 1
        day += timedelta(days=1)
    return counts, not_in_table


# ──────────────────────────────────────────
# xlsx 讀寫（zip 外科手術）
# ──────────────────────────────────────────

def download_xlsx(path):
    subprocess.run([
        GWS, "drive", "files", "get",
        "--params", json.dumps({"fileId": CONTRACT_FILE_ID, "alt": "media"}),
        "--output", path,
    ], check=True)


def upload_xlsx(path):
    subprocess.run([
        GWS, "drive", "files", "update",
        "--params", json.dumps({"fileId": CONTRACT_FILE_ID}),
        "--upload", path,
        "--upload-content-type",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ], check=True)


def read_contract_rows(xlsx_path):
    """用 openpyxl 讀合約名單，回傳符合教練且狀況=進行中的列。

    回傳：[{"row": 列號, "student": 名字, "remain": 剩餘堂數}]
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[_sheet_index_from_xml()]]
    rows = []
    for r in range(1, ws.max_row + 1):
        trainer = ws[f"{COL_TRAINER}{r}"].value
        status  = ws[f"{COL_STATUS}{r}"].value
        student = ws[f"{COL_STUDENT}{r}"].value
        remain  = ws[f"{COL_REMAIN}{r}"].value
        if trainer != TARGET_TRAINER:
            continue
        if status in SKIP_STATUS:
            continue
        if not student:
            continue
        rows.append({"row": r, "student": str(student).strip(),
                     "remain": remain})
    wb.close()
    return rows


def _sheet_index_from_xml():
    """從 CONTRACT_SHEET_XML（sheet4.xml）推回 0-based 分頁 index。"""
    m = re.search(r"sheet(\d+)\.xml", CONTRACT_SHEET_XML)
    return int(m.group(1)) - 1 if m else 0


def patch_remain_cell(xlsx_path, row, new_value):
    """把 xlsx 當 zip，只改 CONTRACT_SHEET_XML 內 R{row} 的數值，其餘原樣複製。"""
    target = f"{COL_REMAIN}{row}"
    tmp = xlsx_path + ".tmp"
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                content = zin.read(item.filename)
                if item.filename == CONTRACT_SHEET_XML:
                    text = content.decode("utf-8")
                    # style code 可能是 s="44" 或 s="61"，用萬用比對
                    pattern = rf'(<c r="{target}"[^>]*><v>)[^<]*(</v></c>)'
                    new_text, n = re.subn(pattern, rf'\g<1>{new_value}\g<2>',
                                          text)
                    if n == 0:
                        raise RuntimeError(
                            f"找不到儲存格 {target}，可能列號位移或該格無值")
                    content = new_text.encode("utf-8")
                zout.writestr(item, content)
    os.remove(xlsx_path)        # rm 被權限擋時用 os.remove
    os.rename(tmp, xlsx_path)


# ──────────────────────────────────────────
# 上次銷課日期
# ──────────────────────────────────────────

def load_last_date():
    if LAST_DATE_FILE.exists():
        data = json.loads(LAST_DATE_FILE.read_text())
        return datetime.strptime(data["last_date"], "%Y-%m-%d").date()
    return None


def save_last_date(d):
    LAST_DATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    LAST_DATE_FILE.write_text(json.dumps({"last_date": d.isoformat()}))


# ──────────────────────────────────────────
# 主流程
# ──────────────────────────────────────────

def main():
    dry_run = "--dry-run" in sys.argv
    today = date.today()

    # 起算日
    since = None
    if "--since" in sys.argv:
        since = datetime.strptime(
            sys.argv[sys.argv.index("--since") + 1], "%Y/%m/%d").date()
    else:
        last = load_last_date()
        if last is None:
            print("找不到上次銷課紀錄，請用 --since YYYY/M/D 指定起算日")
            return
        since = last + timedelta(days=1)

    if since > today:
        print(f"起算日 {since} 晚於今天 {today}，無課可銷。")
        return

    print(f"銷課區間：{since} ~ {today}")
    counts, _ = collect_deductions(since, today)
    if not counts:
        print("區間內沒有要銷的課。")
        return

    # 下載合約單、比對列號
    work = str(Path.home() / "Downloads" / "_contract_work.xlsx")
    download_xlsx(work)
    rows = read_contract_rows(work)
    by_name = {r["student"]: r for r in rows}

    print("\n=== 銷課彙總（尚未寫入）===")
    plan = []
    for name, n in counts.items():
        rec = by_name.get(name)
        if not rec:
            print(f"  ⚠️ {name}：扣 {n} 堂，但合約表查無此人 → 請補登")
            continue
        old = rec["remain"] or 0
        new = old - n
        plan.append((rec["row"], name, old, new))
        flag = "  ⚠️ 已歸零，請確認學員狀態" if new <= 0 else ""
        print(f"  {name}：{old} → {new}（扣 {n}）{flag}")

    if dry_run:
        print("\n[dry-run] 不寫入。確認無誤後拿掉 --dry-run 再跑一次。")
        os.remove(work)
        return

    # 等待人工確認（範例用 input；正式可改成由 Claude 確認）
    if input("\n確認寫入？(yes/no) ").strip().lower() != "yes":
        print("已取消。")
        os.remove(work)
        return

    for row, name, old, new in plan:
        patch_remain_cell(work, row, new)
        print(f"  已改 {name} R{row} = {new}")

    # 讀回驗證
    verify = {r["student"]: r["remain"] for r in read_contract_rows(work)}
    for row, name, old, new in plan:
        if verify.get(name) != new:
            print(f"  ⚠️ 驗證失敗：{name} 應為 {new}，實際 {verify.get(name)}")

    upload_xlsx(work)
    os.remove(work)
    save_last_date(today)

    # 剩餘 ≤ 2 堂提醒
    low = [(n, v) for n, v in verify.items() if isinstance(v, (int, float))
           and v <= 2]
    if low:
        print("\n=== 剩餘 ≤ 2 堂，提醒約續約 ===")
        for n, v in sorted(low, key=lambda x: x[1]):
            print(f"  {n}：剩 {int(v)} 堂")

    print(f"\n完成。已更新上次銷課日期為 {today}。")


if __name__ == "__main__":
    main()
